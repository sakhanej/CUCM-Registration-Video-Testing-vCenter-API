import requests
from requests.auth import HTTPBasicAuth
import xml.etree.ElementTree as ET
import time
import openpyxl
import string
from collections import OrderedDict
import datetime

class EndpointTestingException(Exception):
    def __init__(self, value):
        self.value = value
    def __str__(self):
        return repr(self.value)

def index_to_col(index):
    return string.uppercase[index]

def excel_to_dict(excel_path,sheet_name, headers=[]):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    result_dict = []
    for row in range(2, sheet.max_row +1):
        line = dict()
        for header in headers:
            cell_value = sheet[index_to_col(headers.index(header)) + str(row)].value
            # print cell_value
            if type(cell_value) is unicode:
                cell_value = cell_value.encode('utf-8')
                cell_value = cell_value.strip()
            elif type(cell_value) is int:
                cell_value = str(cell_value)
            elif cell_value is None:
                cell_value = ''
            line[header] = cell_value
        result_dict.append(line)
    wb.close()
    return result_dict

def testing(test_endpoints,production_endpoints,directorylookupstring='',timeout=2):
    results = OrderedDict()
    for production_endpoint in production_endpoints:
        if production_endpoint['Name']!='':
            result = OrderedDict()
            print 'Testing for '+production_endpoint['Name']
            other_stats = other_status(production_endpoint)
            # change the string below to change the directory lookup
            result['NTP Status'] = str(other_stats)
            result['Software Version'] = str(other_stats)
            if isinstance(other_stats,Exception)!=True:
                stats = phonebooksearch(production_endpoint, directorylookupstring, other_stats)
                result = OrderedDict(stats)
                result['Status']='Tested'
                # print result
                for test_endpoint in test_endpoints:
                    if test_endpoint['Name'] != '':
                        if test_endpoint['IP Address'] != '':
                            # production to test
                            if (production_endpoint['Auto Answer']=='Off') & (test_endpoint['Auto Answer']=='Off'):
                                call_stats = connect_endpoints(production_endpoint, test_endpoint, timeout)
                            elif (production_endpoint['Auto Answer']=='On') & (test_endpoint['Auto Answer']=='On'):
                                call_stats = placecall(production_endpoint, test_endpoint, timeout)
                            elif (production_endpoint['Auto Answer']=='On') & (test_endpoint['Auto Answer']=='Off'):
                                call_stats = connect_endpoints(production_endpoint, test_endpoint, timeout)
                            elif (production_endpoint['Auto Answer']=='Off') & (test_endpoint['Auto Answer']=='On'):
                                call_stats = placecall(production_endpoint, test_endpoint, timeout)
                            if isinstance(call_stats,Exception)!=True:
                                if call_stats['Status'] == 'Connected':
                                    result['Call to ' + test_endpoint['Name']] = 'Pass'
                                    result['Call to ' + test_endpoint['Name'] + ' - Output'] = str(call_stats)
                                else:
                                    result['Call to ' + test_endpoint['Name']] = 'Fail'
                                    result['Call to ' + test_endpoint['Name'] + ' - Output'] = str(call_stats)
                            else:
                                result['Call to ' + test_endpoint['Name']] = 'Fail'
                                result['Call to ' + test_endpoint['Name'] + ' - Output'] = str(call_stats)
                            # print call_stats
                            # print result
                            # test to production
                            if (production_endpoint['Auto Answer']=='Off') & (test_endpoint['Auto Answer']=='Off'):
                                call_stats = connect_endpoints(test_endpoint, production_endpoint, timeout)
                            elif (production_endpoint['Auto Answer']=='On') & (test_endpoint['Auto Answer']=='On'):
                                call_stats = placecall(test_endpoint, production_endpoint, timeout)
                            elif (production_endpoint['Auto Answer']=='On') & (test_endpoint['Auto Answer']=='Off'):
                                call_stats = placecall(test_endpoint, production_endpoint, timeout)
                            elif (production_endpoint['Auto Answer']=='Off') & (test_endpoint['Auto Answer']=='On'):
                                call_stats = connect_endpoints(test_endpoint, production_endpoint, timeout)
                            if isinstance(call_stats,Exception)!=True:
                                if call_stats['Status'] == 'Connected':
                                    result['Call from ' + test_endpoint['Name']] = 'Pass'
                                    result['Call from ' + test_endpoint['Name'] + ' - Output'] = str(call_stats)
                                else:
                                    result['Call from ' + test_endpoint['Name']] = 'Fail'
                                    result['Call from ' + test_endpoint['Name'] + ' - Output'] = str(call_stats)
                            else:
                                result['Call from ' + test_endpoint['Name']] = 'Fail'
                                result['Call from ' + test_endpoint['Name'] + ' - Output'] = str(call_stats)
                            # print call_stats
                            # print result
                        elif test_endpoint['IP Address'] == '':
                            call_stats = placecall(production_endpoint, test_endpoint, timeout)
                            if isinstance(call_stats,Exception)!=True:
                                if call_stats['Status'] == 'Connected':
                                    result['Call to ' + test_endpoint['Name']] = 'Pass'
                                    result['Call to ' + test_endpoint['Name'] + ' - Output'] = str(call_stats)
                                else:
                                    result['Call to ' + test_endpoint['Name']] = 'Fail'
                                    result['Call to ' + test_endpoint['Name'] + ' - Output'] = str(call_stats)
                            else:
                                result['Call to ' + test_endpoint['Name']] = 'Fail'
                                result['Call to ' + test_endpoint['Name'] + ' - Output'] = str(call_stats)
                            # print call_stats
                            # print result
            else:
                result['Status'] = 'Not tested'
            results[str(production_endpoint['Name'])] = result
    # print results
    return results

def other_status(production_endpoint):
    other_stats = dict()
    calling_party_get_status = 'http://' + production_endpoint['IP Address'] + '/status.xml'
    try:
        calling_party_status = requests.get(calling_party_get_status,
                                            auth=HTTPBasicAuth(production_endpoint['Username'], production_endpoint['Password']),
                                            headers={'content-type': "application/xml"}, verify=False)
        calling_party_statusXml = calling_party_status.text
        if calling_party_status.status_code==200:
            root = ET.fromstring(calling_party_statusXml)
            other_stats['Software Version'] = root.find('./SystemUnit/Software/Version').text
            other_stats['NTP Status'] = root.find('./NetworkServices/NTP/Status').text
            return other_stats
        elif calling_party_status.status_code==401:
            raise EndpointTestingException('Incorrect Credentials')
    except (requests.exceptions.RequestException,EndpointTestingException) as exception:
        # print 'Exception occured in',ip
        print 'other_status function exception : ' + str(exception)
        return exception

def phonebooksearch(production_endpoint,SearchString,other_stats):
    production_endpoint_put = 'http://' + production_endpoint['IP Address'] + '/putxml'
    phonebooksearchxml = '<Command><Phonebook><Search><SearchField>Name</SearchField><SearchString>'+SearchString+'</SearchString><PhonebookType>Corporate</PhonebookType></Search></Phonebook></Command>'
    try:
        phonebooksearch = requests.request("POST", production_endpoint_put,
                                           auth=HTTPBasicAuth(production_endpoint['Username'],
                                                              production_endpoint['Password']),
                                           data=phonebooksearchxml, headers={'content-type': "application/xml"},
                                           verify=False)
        if phonebooksearch.status_code == 200:
            phonebooksearchresponse = phonebooksearch.text
            # print phonebooksearchresponse
            root = ET.fromstring(phonebooksearchresponse)
            if len(root.findall('./PhonebookSearchResult/Contact')) > 0:
                other_stats['Directory Lookup'] = 'Pass'
            else:
                other_stats['Directory Lookup'] = 'Fail'
            return other_stats
        elif phonebooksearch.status_code == 401:
            raise EndpointTestingException('Incorrect Credentials')
    except (requests.exceptions.RequestException, EndpointTestingException) as exception:
        # print 'Exception occured in',ip
        print 'phonebooksearch function exception : ' + str(exception)
        return exception

def connect_endpoints(calling_party, called_party, timeout):
    try :
        # place call
        print 'Placing call from '+calling_party['Name']+' to '+called_party['Name']
        connectCallrepsonse = connectCall(calling_party,called_party)
        if isinstance(connectCallrepsonse,Exception):
            raise EndpointTestingException('Unable to Dial out from '+calling_party['Name'])
        print 'Timing out for ' + str(timeout)+' second'
        time.sleep(timeout)
        #read call id
        print 'Reading call-id from ' + called_party['Name']
        call_id = readCallId(called_party)
        if isinstance(call_id,Exception):
            raise EndpointTestingException('Unable to read call id from '+called_party['Name'])
        # recieve call based on call id
        print 'Recieving call from ' + called_party['Name']
        recieveCallrepsonse = recieveCall(called_party, call_id)
        if isinstance(recieveCallrepsonse, Exception):
            raise EndpointTestingException('Unable to Recieve call from from ' + called_party['Name'])
        # read status to collect call information
        print 'Reading call statisctics from ' + calling_party['Name']
        call_stats = readCallStatus(calling_party)
        if isinstance(call_stats,Exception):
            raise EndpointTestingException('Unable to read call statisctics from '+calling_party['Name'])
        # disconnect the call
        print 'Disconnecting call from ' + called_party['Name']
        disconnectcallresponse = disconnectcall(called_party, call_id)
        if isinstance(disconnectcallresponse,Exception):
            raise EndpointTestingException('Unable to disconnect call from '+called_party['Name'])
        return call_stats
    except EndpointTestingException as exception:
        print 'connect_endpoints function exception : ' + str(exception)
        return exception

def placecall(calling_party,called_party,timeout):
    try:
        # place call
        print 'Placing call from ' + calling_party['Name'] + ' to ' + called_party['Name']
        connectCallrepsonse = connectCall(calling_party, called_party)
        if isinstance(connectCallrepsonse,Exception):
            raise EndpointTestingException('Unable to Dial out from '+calling_party['Name'])
        print 'Timing out for '+str(timeout)+' second'
        time.sleep(timeout)
        # read call id
        print 'Reading call-id from ' + calling_party['Name']
        call_id = readCallId(calling_party)
        if isinstance(call_id,Exception):
            raise EndpointTestingException('Unable to read call id from '+called_party['Name'])
        # read status to collect call information
        print 'Reading call statisctics from ' + calling_party['Name']
        call_stats = readCallStatus(calling_party)
        if isinstance(call_stats,Exception):
            raise EndpointTestingException('Unable to read call statisctics from ' + calling_party['Name'])
        # disconnect the call
        print 'Disconnecting call from ' + calling_party['Name']
        disconnectcallresponse = disconnectcall(calling_party, call_id)
        if isinstance(disconnectcallresponse,Exception):
            raise EndpointTestingException('Unable to disconnect call from '+called_party['Name'])
        return call_stats
    except EndpointTestingException as exception:
        print 'placecall function exception : ' + str(exception)
        return exception

def connectCall(calling_party,called_party):
    calling_party_put = 'http://' + calling_party['IP Address'] + '/putxml'
    placecallxml = '<Command><Dial><Number>' + str(called_party['URI']) + '</Number></Dial></Command>'
    # print calling_party_put,placecallxml
    try:
        placecall = requests.request("POST", calling_party_put, auth=HTTPBasicAuth(calling_party['Username'], calling_party['Password']), data=placecallxml,headers={'content-type': "application/xml"},verify=False)
        # print placecall.text
        return placecall.text
        if placecall.status_code == 401:
            raise EndpointTestingException('Incorrect Credentials for'+calling_party['Name'])
    except (requests.exceptions.RequestException, EndpointTestingException) as exception:
        # print 'Exception occured in',ip
        print 'connectCall function exception : ' + str(exception)
        return exception

def readCallId(called_party):
    called_party_get_status = 'http://' + called_party['IP Address'] + '/status.xml'
    try:
        called_party_status = requests.get(called_party_get_status,auth=HTTPBasicAuth(called_party['Username'], called_party['Password']),headers={'content-type': "application/xml"},verify=False)

        if called_party_status.status_code == 200:
            called_party_statusXml = called_party_status.text
            root = ET.fromstring(called_party_statusXml)
            # print called_party_statusXml
            call_tag = root.find('./Call')
            if call_tag != None:
                call_id = call_tag.attrib['item']
                return call_id
            else:
                raise EndpointTestingException('Unable to read call id from ' + called_party['Name'])
        elif called_party_status.status_code == 401:
            raise EndpointTestingException('Incorrect Credentials for' + called_party['Name'])
    except (requests.exceptions.RequestException, EndpointTestingException) as exception:
        # print 'Exception occured in',ip
        print 'readCallId function exception : ' + str(exception)
        return exception

def recieveCall(called_party,call_id):
    called_party_put = 'http://' + called_party['IP Address'] + '/putxml'
    recieve_call_xml = '<Command><Call><Accept>' + call_id + '</Accept></Call></Command>'
    try:
        recieve_call = requests.request("POST", called_party_put,
                                        auth=HTTPBasicAuth(called_party['Username'], called_party['Password']),
                                        data=recieve_call_xml, headers={'content-type': "application/xml"},
                                        verify=False)
        return recieve_call.text
        if recieve_call.status_code == 401:
            raise EndpointTestingException('Incorrect Credentials for' + called_party['Name'])
    except (requests.exceptions.RequestException, EndpointTestingException) as exception:
        # print 'Exception occured in',ip
        print 'recieveCall function exception : ' + str(exception)
        return exception

def readCallStatus(calling_party):
    call_stats = dict()
    calling_party_get_status = 'http://' + calling_party['IP Address'] + '/status.xml'
    try:
        calling_party_status = requests.get(calling_party_get_status,auth=HTTPBasicAuth(calling_party['Username'], calling_party['Password']),headers={'content-type': "application/xml"},verify=False)
        # print calling_party_statusXml
        if calling_party_status.status_code == 200:
            calling_party_statusXml = calling_party_status.text
            root = ET.fromstring(calling_party_statusXml)
            call_stats['Status'] = root.find('./Call/Status').text
            call_stats['ReceiveCallRate'] = root.find('./Call/ReceiveCallRate').text
            call_stats['TransmitCallRate'] = root.find('./Call/TransmitCallRate').text
            media_channels = root.findall('./MediaChannels/Call/Channel')
            for media_channel in media_channels:
                type_of_channel = media_channel.find('Type').text
                direction_channel = media_channel.find('Direction').text
                if type_of_channel == 'Audio':
                    call_stats[type_of_channel + '_' + direction_channel] = media_channel.find(
                        type_of_channel + '/Protocol').text
                elif type_of_channel == 'Video':
                    channel_role_tag = media_channel.find(type_of_channel + '/ChannelRole')
                    channel_role = channel_role_tag.text
                    call_stats[channel_role + '_' + type_of_channel + '_' + direction_channel] = media_channel.find(
                        type_of_channel + '/Protocol').text
                else:
                    call_stats[type_of_channel + '_' + direction_channel] = 'True'
            call_stats['MediaChannnelQuantity'] = len(root.findall('./MediaChannels/Call/Channel'))
            return call_stats
        elif calling_party_status.status_code == 401:
            raise EndpointTestingException('Incorrect Credentials for' + calling_party['Name'])
    except (requests.exceptions.RequestException, EndpointTestingException) as exception:
        # print 'Exception occured in',ip
        print 'readCallStatus function exception : ' + str(exception)
        return exception

def disconnectcall(called_party,call_id):
    called_party_put = 'http://' + called_party['IP Address'] + '/putxml'
    disconnect_call_xml = '<Command><Call><Disconnect>' + call_id + '</Disconnect></Call></Command>'
    try:
        disconnect_call = requests.request("POST", called_party_put,
                                           auth=HTTPBasicAuth(called_party['Username'], called_party['Password']),
                                           data=disconnect_call_xml, headers={'content-type': "application/xml"},
                                           verify=False)
        return disconnect_call.text
        if disconnect_call.status_code == 401:
            raise EndpointTestingException('Incorrect Credentials for' + called_party['Name'])
    except (requests.exceptions.RequestException, EndpointTestingException) as exception:
        # print 'Exception occured in',ip
        print 'disconnectcall function exception : ' + str(exception)
        return exception

def write_results_to_excel(results,headers,input_filepath,input_filename):
    value = results[results.keys()[0]]
    new_headers = value.keys()
    # print new_headers
    wb = openpyxl.load_workbook(input_filepath+'\\'+input_filename+'.xlsx')
    sheet = wb['Production Endpoints']
    # Writing Values
    for row in range(1, sheet.max_row + 1):
        # Writing headers
        if row == 1:
            for header in new_headers:
                sheet[index_to_col(new_headers.index(header) + len(headers)) + str(row)] = header
        else:
            cell_value = sheet['A' + str(row)].value
            if type(cell_value) is unicode:
                cell_value = cell_value.encode('utf-8')
                cell_value = cell_value.strip()
            elif type(cell_value) is int:
                cell_value = str(cell_value)
            elif cell_value is None:
                cell_value = ''
            if cell_value != '':
                result = results[str(cell_value)]
                for header in new_headers:
                    if header in result.keys():
                        sheet[index_to_col(new_headers.index(header) + len(headers)) + str(row)] = result[header]
                    else:
                        sheet[index_to_col(new_headers.index(header) + len(headers)) + str(row)] = ''
    ts = time.time()
    st = datetime.datetime.fromtimestamp(ts).strftime('%d%m%Y_%H%M%S')
    print 'Testing completed, results are stored in '+input_filepath+'\\'+input_filename+ '_output_' + st + '.xlsx'
    wb.save(input_filepath+'\\'+input_filename+ '_output_' + st + '.xlsx')
    wb.close()

def main():
    headers = ['Name','IP Address','Username','Password','URI','Auto Answer']
    input_filepath = raw_input("Enter the file Path of the Input File :")
    input_filename = raw_input("Enter file name (without .xlsx) :")
    lookupstring = str(raw_input("[Optional] Enter the Value for Corporate Directory lookup :"))
    input_timeout = str(raw_input("[Optional] Enter the time for script to wait in order to recieve call from the Called party (in seconds)[Default = 4]"))
    if lookupstring == '':
        lookupstring = 'Singapore Phone 3'
    if input_timeout == '':
        input_timeout = 4
    else:
        input_timeout = int(input_timeout)
    test_endpoints = excel_to_dict(input_filepath+'\\'+input_filename+'.xlsx','Test Endpoints',headers)
    production_endpoints = excel_to_dict(input_filepath+'\\'+input_filename+'.xlsx','Production Endpoints',headers)
    results = testing(test_endpoints,production_endpoints,directorylookupstring=lookupstring,timeout=input_timeout)
    write_results_to_excel(results,headers,input_filepath,input_filename)

main()
